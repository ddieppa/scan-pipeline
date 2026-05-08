from __future__ import annotations

from pathlib import Path

from app.extractors.common import ExtractionResult


def extract_xlsx(path: Path, max_sheets: int, max_rows_per_sheet: int) -> ExtractionResult:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX extraction") from exc

    workbook = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        chunks: list[str] = []
        sheet_names = workbook.sheetnames[:max_sheets]
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            chunks.append(f"[Sheet] {sheet_name}")
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index >= max_rows_per_sheet:
                    break
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    chunks.append(" | ".join(values))
        return ExtractionResult(
            text="\n".join(chunks),
            text_source="xlsx",
            pages_inspected=len(sheet_names),
            needs_ocr=False,
            metadata={"sheet_count": len(workbook.sheetnames)},
        )
    finally:
        workbook.close()
